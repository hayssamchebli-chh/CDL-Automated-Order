import io
import os
import re
import math
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from datetime import datetime
import tempfile

# =========================
# LOAD STATIC ASSETS (Mode B)
# =========================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

MIN_PATH = os.path.join(BASE_DIR, "assets", "Minimum Threshold.xlsx")
TEMPLATE_PATH = os.path.join(BASE_DIR, "assets", "Order Template.xlsm")

with open(MIN_PATH, "rb") as f:
    min_bytes = f.read()

with open(TEMPLATE_PATH, "rb") as f:
    template_bytes = f.read()




# =========================
# Business rules lists
# =========================
MULTIPLE_15 = {
    "CDL-NYA 3-029 WT", "CDL-NYA 3-029 BL", "CDL-NYA 3-029 GY",
    "CDL-NYA 3-029 YL", "CDL-NYA 3-029 BK", "CDL-NYA 3-029 RD",
    "CDL-NYA 3-029 GN", "CDL-NYA 3-029 GN-YL",
    "CDL-NYA 1.5 WT", "CDL-NYA 1.5 BL", "CDL-NYA 1.5 GY",
    "CDL-NYA 1.5 YL", "CDL-NYA 1.5 BK", "CDL-NYA 1.5 RD",
    "CDL-NYA 1.5 GN", "CDL-NYA 1.5 GN-YL",
    "CDL-NYAF 1.5 BK", "CDL-NYAF 1.5 RD"
}

MULTIPLE_10 = {
    "CDL-NYAF 2.5 BK", "CDL-NYAF 2.5 RD",
    "CDL-NYZ 2X0.5 RN", "CDL-NYZ 2X1", "CDL-NYZ 2X1 RN",
    "CDL-NYZ 2X1.5", "CDL-NYZ 2X1.5 RN",
    "CDL-NYZ 2X2", "CDL-NYZ 2X2 RN",
    "CDL-NYZ 2X2.5", 
    "CDL-NYA 2 WT","CDL-NYA 2 WT",	"CDL-NYA 2 WT",	"CDL-NYA 2 WT",	"CDL-NYA 2 WT",	"CDL-NYA 2 WT",	
    "CDL-NYA 2 WT",	"CDL-NYA 2 WT"
}

COLOR_HEADERS = ["B", "BE", "G", "J", "N", "R", "V", "VJ", "RN"]

# =========================
# Helpers
# =========================
def normalize_cell(x) -> str:
    return re.sub(r"\s+", " ", str(x).strip()) if x is not None else ""

def read_excel_auto_header_bytes(file_bytes: bytes, required_columns, sheet_name=0) -> pd.DataFrame:
    """
    Read Excel from bytes and auto-detect the header row by searching for required_columns.
    """
    # read without header
    raw = pd.read_excel(io.BytesIO(file_bytes), header=None, sheet_name=sheet_name)

    header_row = None
    for i, row in raw.iterrows():
        row_vals = row.astype(str).str.strip().values
        if all(col in row_vals for col in required_columns):
            header_row = i
            break

    if header_row is None:
        raise ValueError(f"Could not auto-detect header row. Expected columns: {required_columns}")

    df = pd.read_excel(io.BytesIO(file_bytes), header=header_row, sheet_name=sheet_name)
    df.columns = df.columns.astype(str).str.strip()
    return df

def validate_minimum_mapping(min_df: pd.DataFrame) -> pd.DataFrame:
    """
    Ensures Minimum file contains required columns and normalizes them.
    """
    min_df.columns = min_df.columns.astype(str).str.strip()
    if "Minimum" not in min_df.columns and "minimum" in min_df.columns:
        min_df = min_df.rename(columns={"minimum": "Minimum"})

    required = ["Item No.1", "Minimum", "Template TYPE", "Template SIZE", "Template Color"]
    missing = [c for c in required if c not in min_df.columns]
    if missing:
        raise ValueError(f"Minimum file is missing required columns: {missing}")

    # clean text columns
    for c in ["Item No.1", "Template TYPE", "Template SIZE", "Template Color"]:
        min_df[c] = min_df[c].astype(str).str.strip()

    # numeric minimum
    min_df["Minimum"] = pd.to_numeric(min_df["Minimum"], errors="coerce")
    return min_df

def build_order_df(stock_df: pd.DataFrame, min_df: pd.DataFrame) -> pd.DataFrame:
    """
    Merge stock+minimum, apply rules, produce order_df (with Order Qty).
    """
    # normalize keys
    stock_df["Item No.1"] = stock_df["Item No.1"].astype(str).str.strip()
    min_df["Item No.1"]   = min_df["Item No.1"].astype(str).str.strip()

    # numeric available
    stock_df["Stock Available Quantity"] = pd.to_numeric(
        stock_df["Stock Available Quantity"], errors="coerce"
    ).fillna(0)

    # merge
    df = stock_df.merge(
        min_df[["Item No.1", "Minimum", "Template TYPE", "Template SIZE", "Template Color"]],
        on="Item No.1",
        how="inner",
    )

    # drop missing minimum if any
    df = df.dropna(subset=["Minimum", "Stock Available Quantity"])

    order_list = []
    for _, row in df.iterrows():
        item = str(row["Item No.1"]).strip()
        available = float(row["Stock Available Quantity"])
        minimum = float(row["Minimum"])

        shortage = minimum - available
        if pd.isna(shortage) or shortage <= 0:
            continue

        # Tolerance rule: if Minimum > 20, ignore if shortage <= 3
        if minimum > 20 and shortage <= 3:
            continue

        # Rounding rules
        if item in MULTIPLE_15:
            order_qty = math.ceil(shortage / 15) * 15
        elif item in MULTIPLE_10:
            order_qty = math.ceil(shortage / 10) * 10
        else:
            order_qty = math.ceil(shortage)

        order_list.append({
            "Item No.1": item,
            "Available": available,
            "Minimum": minimum,
            "Shortage": shortage,
            "Order Qty": order_qty,
            "Template TYPE": row["Template TYPE"],
            "Template SIZE": row["Template SIZE"],
            "Template Color": row["Template Color"],
        })

    order_df = pd.DataFrame(order_list)
    if not order_df.empty:
        order_df = order_df.sort_values(by=["Item No.1"]).reset_index(drop=True)

    # validate mapping is present for all rows
    missing_map = order_df[
        order_df["Template TYPE"].isna()
        | order_df["Template SIZE"].isna()
        | order_df["Template Color"].isna()
    ]
    if not missing_map.empty:
        bad = missing_map[["Item No.1"]].head(50).to_string(index=False)
        raise ValueError("Some ordered items are missing mapping columns in Minimum file.\n" + bad)

    # validate Template Color values
    bad_colors = order_df[~order_df["Template Color"].isin(COLOR_HEADERS)]
    if not bad_colors.empty:
        bad = bad_colors[["Item No.1", "Template Color"]].head(50).to_string(index=False)
        raise ValueError(
            "Some items have Template Color not in the template headers (B,BE,G,J,N,R,V,VJ,RN).\n" + bad
        )

    return order_df

def find_header_row(ws, must_have=("TYPE", "SIZE"), scan_rows=120):
    must_have = set(must_have)
    for r in range(1, min(ws.max_row, scan_rows) + 1):
        row_vals = {normalize_cell(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)}
        if must_have.issubset(row_vals):
            return r
    return None

def build_col_map(ws, header_row):
    m = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is None:
            continue
        key = normalize_cell(v)
        if key:
            m[key] = c
    return m

def fill_template_xlsm_bytes(
    template_xlsm_bytes: bytes,
    order_df: pd.DataFrame,
    sheet_name: str = "Main Pricing Template",
    stop_after_blank_type_rows: int = 25
) -> bytes:
    """
    Load XLSM (keep VBA), clear all color cells, fill new values, return XLSM bytes.
    """
    # openpyxl needs a file path for keep_vba to behave reliably
    with tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False) as tmp_in:
        tmp_in.write(template_xlsm_bytes)
        tmp_in.flush()
        in_path = tmp_in.name

    wb = load_workbook(in_path, keep_vba=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]

    header_row = find_header_row(ws, must_have=("TYPE", "SIZE"))
    if header_row is None:
        raise ValueError("Could not find a row containing headers TYPE and SIZE in the template.")

    col_map = build_col_map(ws, header_row)

    type_col = col_map.get("TYPE")
    size_col = col_map.get("SIZE")
    if not type_col or not size_col:
        raise ValueError("Found header row, but could not map TYPE/SIZE columns.")

    color_cols = {h: col_map.get(h) for h in COLOR_HEADERS}
    missing = [h for h, c in color_cols.items() if c is None]
    if missing:
        raise ValueError(f"Template is missing these color headers: {missing}")

    # CLEAR color cells
    blank_type_streak = 0
    for r in range(header_row + 1, ws.max_row + 1):
        t = normalize_cell(ws.cell(r, type_col).value)
        if not t:
            blank_type_streak += 1
            if blank_type_streak >= stop_after_blank_type_rows:
                break
        else:
            blank_type_streak = 0

        for h in COLOR_HEADERS:
            ws.cell(r, color_cols[h]).value = None

    # Build lookup (TYPE, SIZE) -> row
    row_lookup = {}
    blank_type_streak = 0
    for r in range(header_row + 1, ws.max_row + 1):
        t = normalize_cell(ws.cell(r, type_col).value)
        s = normalize_cell(ws.cell(r, size_col).value)

        if not t:
            blank_type_streak += 1
            if blank_type_streak >= stop_after_blank_type_rows:
                break
            continue
        blank_type_streak = 0
        row_lookup[(t, s)] = r

    # Fill
    not_found = []
    for _, rec in order_df.iterrows():
        template_type = normalize_cell(rec["Template TYPE"])
        template_size = normalize_cell(rec["Template SIZE"])
        color_code    = normalize_cell(rec["Template Color"])
        qty           = float(rec["Order Qty"])

        target_row = row_lookup.get((template_type, template_size))
        if target_row is None:
            not_found.append((rec["Item No.1"], template_type, template_size))
            continue

        ws.cell(target_row, color_cols[color_code]).value = qty

    # Save to bytes
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    if not_found:
        # Not fatal, but useful to show
        st.warning(f"{len(not_found)} items could not be matched to a row in the template. "
                   f"Example: {not_found[:5]}")

    return out.getvalue()

def df_to_excel_bytes(df: pd.DataFrame, sheet_name="Order") -> bytes:
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    out.seek(0)
    return out.getvalue()

# =========================
# Streamlit UI
# =========================
st.set_page_config(page_title="Stock Replenishment Generator", layout="wide")
st.title("Stock Replenishment Generator")

st.write("Upload the **Stock Report**. Then generate the order list and the filled order template (.xlsm).")

stock_file = st.file_uploader("Upload Stock Report", type=["xlsx", "xls"])


sheet_name = st.text_input("Template sheet name", value="Main Pricing Template")
st.caption("If your template uses another sheet name, change it here.")

if st.button("Generate Order", type="primary", disabled=not stock_file):
    try:
        with st.spinner("Reading files..."):
            stock_bytes = stock_file.getvalue()


            stock_df = read_excel_auto_header_bytes(stock_bytes, ["Item No.1", "Stock Available Quantity"])
            min_df_raw = read_excel_auto_header_bytes(min_bytes, ["Item No.1"])  # header auto-detect
            min_df = validate_minimum_mapping(min_df_raw)

        with st.spinner("Computing replenishment quantities..."):
            order_df = build_order_df(stock_df, min_df)

        st.success(f"Computed {len(order_df)} order line(s).")
        st.dataframe(order_df, use_container_width=True)

        # Outputs
        stamp = datetime.now().strftime("%Y-%m-%d_%H%M")
        order_list_bytes = df_to_excel_bytes(order_df, sheet_name="Generated_Order")

        with st.spinner("Filling XLSM template (clearing + overwriting matrix)..."):
            filled_xlsm_bytes = fill_template_xlsm_bytes(
                template_xlsm_bytes=template_bytes,
                order_df=order_df,
                sheet_name=sheet_name,
                stop_after_blank_type_rows=25
            )

        col1, col2 = st.columns(2)
        with col1:
            st.download_button(
                label="Download Generated_Order.xlsx",
                data=order_list_bytes,
                file_name=f"Generated_Order_{stamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        with col2:
            st.download_button(
                label="Download Order Template - FILLED.xlsm",
                data=filled_xlsm_bytes,
                file_name=f"Order_Template_FILLED_{stamp}.xlsm",
                mime="application/vnd.ms-excel.sheet.macroEnabled.12"
            )

    except Exception as e:
        st.error(str(e))

        st.exception(e)






